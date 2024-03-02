import openpyxl
import random
import time
from datetime import datetime

# Create a new Excel workbook
workbook = openpyxl.Workbook()
sheet = workbook.active

# Set column headers
sheet['A1'] = 'Time'  # Time column header
sheet['B1'] = 'A1'
sheet['C1'] = 'B1'
sheet['D1'] = 'C1'
sheet['E1'] = 'D1'

# Function to generate random number and write to Excel, including the time
def generate_and_write():
    row_num = sheet.max_row + 1  # Get the next row number
    # Write the current time in the first column
    current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    sheet.cell(row=row_num, column=1).value = current_time
    print(f"Writing Time to cell {row_num}, 1")
    
    # Generate and write random numbers for the other columns
    for col in range(2, 6):  # Adjusted for the added Time column
        sheet.cell(row=row_num, column=col).value = random.randint(1, 5)  # Adjusted random range to 1-100 for variety
        print(f"Writing to cell {row_num}, {col}")
    workbook.save('data.xlsx')  # Save workbook

# Generate and write random numbers every 1 second
try:
    while True:
        generate_and_write()
        time.sleep(0.2)  # Sleep for 1 second
except KeyboardInterrupt:
    print("Process interrupted.")
